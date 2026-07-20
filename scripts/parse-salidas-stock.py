#!/usr/bin/env python3
"""
Lee los reportes "Salidas de Stocks" que exporta el sistema de facturación
(uno por empresa: Cerámica, Porcelanas, Presupuesto — formato .xlsx real, no
SpreadsheetML) y genera salidas_stock_export.json, que después carga
import-salidas-stock.cjs en `factura_items` (cruzando por número de
comprobante contra las facturas ya cargadas por Importar ventas).

Estructura del archivo (fila 12 = encabezado, desde fila 13 = datos, hasta
una fila "TOTAL" al final):
  Venta | Salida | Cliente | Item | Cantidad | Precio Vta | Costo Unit. | Costo Total | Observaciones

- "Venta" viene como texto combinado: "dd/mm/aa  TIPO  NUMERO" (ej.
  "02/01/26  F A  00011-00001224") — hay que separarlo en fecha/tipo/número.
  Las notas de crédito/débito vienen igual ("NC A", "ND A").
- "Item" viene como "CODIGO  DESCRIPCION" (ej. "0011  INODORO CORTO NAPOLES").
  El código no siempre significa lo mismo entre líneas (el mismo dígito de
  "tipo" se reusa para cosas distintas), así que el mapeo a pieza del catálogo
  se hace por el texto completo del Item (código+descripción), no decodificando
  el número — ver ITEM_A_PIEZA más abajo.

Uso:
    python3 scripts/parse-salidas-stock.py "SalidasStocks Cerámica.xlsx" "SalidasStocks Porcelanas.xlsx" "SalidasStocks Presupuesto.xlsx"
"""
import json, os, re, sys, unicodedata
import openpyxl

OUT = os.path.join(os.path.dirname(__file__), "..", "salidas_stock_export.json")

# Item (código + descripción, tal cual aparece) -> (linea, tipo_pieza, variante, calidad)
# Debe coincidir exactamente con las combinaciones cargadas en scripts/setup-piezas.cjs.
ITEM_A_PIEZA = {
    "0011 INODORO CORTO NAPOLES": ("Napoles", "Inodoro corto", "", "1era"),
    "0011 INODORO NAPOLES": ("Napoles", "Inodoro corto", "", "1era"),
    "0012 INODORO CORTO NAPOLES": ("Napoles", "Inodoro corto", "", "comercial"),
    "0012 INODORO NAPOLES": ("Napoles", "Inodoro corto", "", "comercial"),
    "0013 INODORO CORTO NAPOLES": ("Napoles", "Inodoro corto", "", "3era"),
    "0021 MOCHILA A CODO": ("Napoles", "Deposito de codo", "", "1era"),
    "0022 MOCHILA A CODO": ("Napoles", "Deposito de codo", "", "comercial"),
    "0031 BIDET NAPOLES": ("Napoles", "Bidet", "3 agujeros", "1era"),
    "0032 BIDET NAPOLES": ("Napoles", "Bidet", "3 agujeros", "comercial"),
    "0041 LAVATORIO": ("Napoles", "Lavatorio", "", "1era"),
    "0042 LAVATORIO": ("Napoles", "Lavatorio", "", "comercial"),
    "0051 COLUMNA": ("Napoles", "Columna", "", "1era"),
    "0052 COLUMNA": ("Napoles", "Columna", "", "comercial"),
    "0061 LAVATORIO MONOCOMANDO": ("Napoles", "Lavatorio", "Monocomando", "1era"),
    "0062 LAVATORIO MONOCOMANDO": ("Napoles", "Lavatorio", "Monocomando", "comercial"),
    "0111 INODORO LARGO LYON": ("Lyon", "Inodoro largo", "", "1era"),
    "0112 INODORO LARGO LYON": ("Lyon", "Inodoro largo", "", "comercial"),
    "0113 INODORO LARGO LYON": ("Lyon", "Inodoro largo", "", "3era"),
    "0121 MOCHILA DE APOYO": ("Lyon", "Deposito de apoyo", "", "1era"),
    "0122 MOCHILA DE APOYO": ("Lyon", "Deposito de apoyo", "", "comercial"),
    "0131 BIDET LYON": ("Lyon", "Bidet", "3 agujeros", "1era"),
    "0132 BIDET LYON": ("Lyon", "Bidet", "3 agujeros", "comercial"),
    "0141 BIDET LYON MONOCOMANDO": ("Lyon", "Bidet", "Monocomando", "1era"),
    "0142 BIDET LYON MONOCOMANDO": ("Lyon", "Bidet", "Monocomando", "comercial"),
    "0211 BACHA CANCUN": ("Bachas", "Cancún", "", "1era"),
    "0212 BACHA CANCUN": ("Bachas", "Cancún", "", "comercial"),
    "0311 COMBO LIRA": ("Lira", "Combo (inodoro largo + depósito)", "", "1era"),
    "0312 COMBO LIRA": ("Lira", "Combo (inodoro largo + depósito)", "", "comercial"),
    "0321 COMBO BELMOND": ("Belmond", "Combo (inodoro largo + depósito)", "", "1era"),
    "0322 COMBO BELMOND": ("Belmond", "Combo (inodoro largo + depósito)", "", "comercial"),
    "0411 ELEMENTO DE MOCHILA": ("Repuestos", "Elemento de mochila", "", "comercial"),
    "0421 TAPA DE MOCHILA": ("Repuestos", "Tapa de mochila", "", "comercial"),
    "0431 TAPA DE INODORO FL (NÁPOLES)": ("Repuestos", "Tapa de inodoro", "Napoles", "comercial"),
    "0441 TAPA DE INODORO BR (LYON)": ("Repuestos", "Tapa de inodoro", "Lyon", "comercial"),
    "0451 ELEMENTO DE MOCHILA LIRA/BELMOND": ("Repuestos", "Elemento de mochila", "Lira/Belmond", "comercial"),
    "0461 TAPA DE INODORO BELMOND": ("Repuestos", "Tapa de inodoro", "Belmond", "comercial"),
    "999 BACHA BOWL MONOCOMANDO": ("Otros", "Bacha bowl monocomando", "", "comercial"),
    "999 MUEBLE DE VANITORY BLANCO": ("Otros", "Mueble de vanitory", "Blanco", "comercial"),
    "999 MUEBLE DE VANITORY WENGUE": ("Otros", "Mueble de vanitory", "Wengue", "comercial"),
}


def sin_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def detect_empresa(filename):
    low = sin_acentos(filename.lower())
    if "ceram" in low:
        return "Ceramica"
    if "porcelan" in low:
        return "Porcelanas"
    if "presupuesto" in low:
        return "Presupuesto"
    raise ValueError(f"No pude detectar la empresa a partir del nombre de archivo: {filename}")


def parsear_venta(venta_raw):
    # "02/01/26  F A  00011-00001224" -> fecha, tipo_comprobante, numero_comprobante
    partes = venta_raw.split()
    if len(partes) < 3:
        return None
    fecha_raw = partes[0]
    numero = partes[-1]
    tipo = " ".join(partes[1:-1])
    d, m, a = fecha_raw.split("/")
    anio = "20" + a if len(a) == 2 else a
    fecha = f"{anio}-{m.zfill(2)}-{d.zfill(2)}"
    return fecha, tipo, numero


def parsear_archivo(path):
    empresa = detect_empresa(os.path.basename(path))
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    registros = []
    sin_mapear = set()
    for row in ws.iter_rows(min_row=13, values_only=True):
        venta_raw = row[1] if len(row) > 1 else None
        item_raw = row[8] if len(row) > 8 else None
        cantidad = row[10] if len(row) > 10 else None
        precio = row[11] if len(row) > 11 else None
        if not venta_raw or not isinstance(venta_raw, str):
            continue
        if venta_raw.strip().upper().startswith("TOTAL"):
            continue
        parsed = parsear_venta(venta_raw)
        if not parsed:
            continue
        fecha, tipo_comprobante, numero_comprobante = parsed
        item_key = " ".join(str(item_raw).split()) if item_raw else None
        pieza = ITEM_A_PIEZA.get(item_key)
        if not pieza:
            if item_key:
                sin_mapear.add(item_key)
            continue
        linea, tipo_pieza, variante, calidad = pieza
        registros.append({
            "empresa": empresa,
            "fecha": fecha,
            "tipo_comprobante": tipo_comprobante,
            "numero_comprobante": numero_comprobante,
            "linea": linea,
            "tipo_pieza": tipo_pieza,
            "variante": variante,
            "calidad": calidad,
            "cantidad": int(cantidad) if cantidad else 0,
            "precio_vta": float(precio) if precio else None,
        })

    if sin_mapear:
        print(f"  ⚠ {path}: {len(sin_mapear)} ítem(s) sin mapeo en ITEM_A_PIEZA (se saltearon): {sorted(sin_mapear)}")

    return registros


def main():
    paths = sys.argv[1:]
    if not paths:
        print("Uso: python3 scripts/parse-salidas-stock.py archivo1.xlsx [archivo2.xlsx ...]")
        sys.exit(1)

    todos = []
    for path in paths:
        registros = parsear_archivo(path)
        print(f"{os.path.basename(path)}: {len(registros)} renglones leídos ({detect_empresa(path)})")
        todos.extend(registros)

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)
    print(f"\nTotal: {len(todos)} renglones -> {OUT}")


if __name__ == "__main__":
    main()
