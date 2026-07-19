#!/usr/bin/env python3
"""
Lee "Consolidado Ventas_Produccion 2026.xlsx" y genera produccion_export.json
(una fila por fecha + pieza + calidad + tipo), que luego carga import-produccion.cjs.

- Producción y Venta: se leen del DETALLE DIARIO por calidad (1°/2°/3°).
- Rotura: el Excel solo la tiene como total mensual unificado (sin día ni calidad),
  así que se emite un total por mes (último día hábil del mes) mapeado a calidad 'comercial'.

Cada pieza de la planilla se traduce al catálogo del CRM (linea / tipo_pieza / variante / calidad)
para que luego import-produccion.cjs pueda resolver el pieza_id y habilitar los cruces con ventas.
"""
import json, os, sys
from collections import defaultdict
import openpyxl

SRC = os.environ.get(
    "PRODUCCION_XLSX",
    "/Users/victorworonczuk/Downloads/Consolidado Ventas_Produccion 2026.xlsx",
)
OUT = os.path.join(os.path.dirname(__file__), "..", "produccion_export.json")

def norm(s):
    return " ".join(str(s).split()).strip() if s is not None else ""

# Planilla de producción -> catálogo del CRM (linea, tipo_pieza, variante)
PIEZA_MAP = {
    "Inodoro Napoles":        ("Napoles", "Inodoro corto",     ""),
    "Bidet Napoles":          ("Napoles", "Bidet",             "3 agujeros"),
    "Inodoro Lyon":           ("Lyon",    "Inodoro largo",     ""),
    "Bidet Lyon":             ("Lyon",    "Bidet",             "3 agujeros"),
    "Bidet Lyon monocomando": ("Lyon",    "Bidet",             "Monocomando"),
    "Lavatorios":             ("Napoles", "Lavatorio",         ""),
    "Lavatorios Monocomando": ("Napoles", "Lavatorio",         "Monocomando"),
    "Columnas":               ("Napoles", "Columna",           ""),
    "Depositos codo":         ("Napoles", "Deposito de codo",  ""),
    "Depositos apoyo":        ("Lyon",    "Deposito de apoyo", ""),
    "Bacha Cancun":           ("Bachas",  "Cancún",            ""),
}
CAL_MAP = {"1°": "1era", "2°": "comercial", "3°": "3era"}
FILTRO_TIPO = {"Produccion": "produccion", "Venta": "venta"}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    rows = []
    faltantes = set()
    ctrl = defaultdict(int)  # tipo -> suma cantidad (control)

    for month in wb.sheetnames:
        ws = wb[month]
        # Columna TOTAL (varía por mes según días hábiles)
        totcol = None
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(row=4, column=c).value).upper() == "TOTAL":
                totcol = c
                break
        if not totcol:
            print(f"[WARN] {month}: no encontré columna TOTAL, salteo", file=sys.stderr)
            continue
        # Columnas de fecha = 4 (D) .. totcol-1 ; fecha en fila 4
        date_cols = []
        for c in range(4, totcol):
            v = ws.cell(row=4, column=c).value
            if hasattr(v, "year"):
                date_cols.append((c, v.date().isoformat()))
        ultimo_dia = date_cols[-1][1] if date_cols else None

        for r in range(5, ws.max_row + 1):
            p = norm(ws.cell(row=r, column=1).value)
            f = norm(ws.cell(row=r, column=2).value)
            cal = norm(ws.cell(row=r, column=3).value)

            # --- DETALLE DIARIO: Produccion / Venta por calidad ---
            if cal in CAL_MAP and f in FILTRO_TIPO:
                if p not in PIEZA_MAP:
                    faltantes.add(p); continue
                linea, tipo_pieza, variante = PIEZA_MAP[p]
                calidad = CAL_MAP[cal]
                tipo = FILTRO_TIPO[f]
                for c, fecha in date_cols:
                    val = ws.cell(row=r, column=c).value
                    if val:  # solo celdas con cantidad > 0
                        rows.append({
                            "fecha": fecha, "linea": linea, "tipo_pieza": tipo_pieza,
                            "variante": variante, "calidad": calidad, "tipo": tipo,
                            "cantidad": int(val),
                        })
                        ctrl[tipo] += int(val)

            # --- ROTURA: solo total mensual unificado (sin día ni calidad) ---
            elif cal == "Unificado" and f == "Rotura":
                if p not in PIEZA_MAP:
                    faltantes.add(p); continue
                val = ws.cell(row=r, column=totcol).value
                if val and ultimo_dia:
                    linea, tipo_pieza, variante = PIEZA_MAP[p]
                    rows.append({
                        "fecha": ultimo_dia, "linea": linea, "tipo_pieza": tipo_pieza,
                        "variante": variante, "calidad": "comercial", "tipo": "rotura",
                        "cantidad": int(val),
                    })
                    ctrl["rotura"] += int(val)

    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False, indent=0)

    print(f"Filas generadas: {len(rows)}  ->  {os.path.abspath(OUT)}")
    print("Control de totales (deben coincidir con el consolidado anual):")
    for t in ("produccion", "venta", "rotura"):
        print(f"   {t:11s}: {ctrl[t]}")
    if faltantes:
        print("[WARN] piezas sin mapear (revisar PIEZA_MAP):", faltantes, file=sys.stderr)


if __name__ == "__main__":
    main()
